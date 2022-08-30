import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from tm2py.components.component import Component
import string


class ConvergenceReport(Component):
    """Produce Highway assignment convergence metrics.

    Args:
        controller: parent RunController object
    """

    def __init__(self, controller):
        """Constructor for ConvergenceReport components.

        Args:
            controller (RunController): Reference to current run controller.
        """
        super().__init__(controller)
        self.config =  self.controller.config

    def run(self):
        self.export_summaries()
        
        
    def validate_inputs(self):
        pass
        
    def _get_link_attributes(self, network, attributes):
        """Calculate the length attributes used in the highway skims."""
        link_attr_table = []
        for link in network.links():
            link_attr_table.append({
                attr:link[attr] for attr in attributes
            })
        return pd.DataFrame(link_attr_table)
    
    def _delete_sheet_from_xlsx(self, path, name):
        wb = load_workbook(path)
        if name in wb.sheetnames:
            del wb[name]
            wb.save(path)
        wb.close()
            
    
    
    def _create_formula_sheet(self, iteration, main_df):
        
        index_size = main_df.index.nlevels
        num_cols = len(main_df.columns)
        num_rows = len(main_df)
        
        prev_iteration = iteration - 1
        formula_template = "=iter{iteration}!{col}{row} / iter{prev_iteration}!{col}{row} - 1"
    
        formulas = pd.DataFrame(formula_template,
                                  index = range(2, num_rows + 2), 
                                  columns = list(string.ascii_uppercase)[index_size: (index_size + num_cols)]).to_dict()

        for col in formulas:
            for row in formulas[col]:
                 formulas[col][row] = formulas[col][row].format(iteration=iteration, 
                                                                prev_iteration=prev_iteration,
                                                                col=col, 
                                                                row=row)

        formula_df = pd.DataFrame(formulas)
        formula_df.columns = [f"=iter{iteration}!{col}1" for col in formula_df.columns]
        formula_df[' '] = [f"=iter{iteration}!A{row}" for row in formula_df.index]
        formula_df['  '] = [f"=iter{iteration}!B{row}" for row in formula_df.index]
        formula_df = formula_df[sorted(formula_df.columns)]   
        
        return formula_df
        
    
    def export_summaries(self, 
                        network_attrs = ['length','auto_time','auto_volume'], 
                        summary_fields = ['#link_county']):
            
        iteration = self.controller.iteration
        report_path = self.get_abs_path(self.config.highway.convergence.output_convergence_report_path)
        
        if not os.path.exists(report_path):
            pd.DataFrame(['']).to_excel(report_path, sheet_name = 'empty')

        summary_by_time_period = {}
        
        for time in self.time_period_names:
            scenario = self.get_emme_scenario(
                self.config.emme.highway_database_path, time
            )
            network = scenario.get_network()
            attrs = network_attrs + summary_fields
            link_attrs = self._get_link_attributes(network, attrs)
            link_attrs['vmt'] = link_attrs['auto_volume'] * link_attrs['length']
            link_attrs['tt'] = link_attrs['auto_volume'] * link_attrs['auto_time']
            summary = link_attrs.groupby(summary_fields)[['auto_volume','vmt','tt']].sum()
            summary_by_time_period[time] = pd.concat([summary, pd.DataFrame(summary.sum().rename('total')).T])
        
        summary_all = pd.concat(summary_by_time_period, axis = 0)
        summary_all.index.rename(['time_period'] + summary_fields, inplace = True)
        
        self._delete_sheet_from_xlsx(report_path,'Comp with last iter')
        
        with pd.ExcelWriter(report_path, mode="a", engine="openpyxl", if_sheet_exists = 'replace') as writer:
            summary_all.reset_index().to_excel(writer, sheet_name=f'iter{iteration}',index = False)
            if iteration > 0:
                formula_df = self._create_formula_sheet(iteration, summary_all)
                pd.DataFrame(formula_df).to_excel(writer, index = False, sheet_name = 'Comp with last iter')
            writer.save()
        
        self._delete_sheet_from_xlsx(report_path,'empty')
